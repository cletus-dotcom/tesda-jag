from datetime import date, datetime
from io import BytesIO

from app import db
from app.config import (
    CLASSIFICATIONS,
    STATUSES,
    is_valid_action_needed,
    is_valid_doc_type,
    is_valid_rmt_record_type,
    parse_action_fields,
    parse_doc_type_fields,
    parse_recipient_fields,
    parse_rmt_record_type_fields,
    normalize_classification,
    resolve_origin_fields,
)
from app.dept_service import list_department_names
from app.models import DtsDoc, DtsHistory, RmtRecord, log_history
from app.utils import parse_document_date

DTS_IMPORT_FIELDS = [
    ("route_number", "Route Number", True),
    ("classification", "Classification (Incoming/Outgoing)", True),
    ("doc_type", "Document Type", True),
    ("doc_type_part", "Document Type (if Others)", False),
    ("date_received", "Date Received (YYYY-MM-DD)", True),
    ("origin", "Origin", False),
    ("other_office", "Other Office (if Origin is Others)", False),
    ("subject", "Subject", True),
    ("doc_link", "Document Link", False),
    ("action_needed", "Action Needed", False),
    ("action_part", "Action Needed (if Others)", False),
    ("action_particulars", "Action Particulars", False),
    ("resp_unit", "Responsible Unit", False),
    ("responsible_person", "Responsible Person", False),
    ("action_provided", "Action Provided", False),
    ("date_accomp", "Date Accomplished (YYYY-MM-DD)", False),
    ("status", "Status", False),
    ("forwarded_to", "Forwarded To / Submitted To", False),
    ("forwarded_to_part", "Recipient (if Others)", False),
    ("remarks", "Remarks", False),
]

RMT_IMPORT_FIELDS = [
    ("record_number", "Record Number", True),
    ("title", "Title", True),
    ("record_type", "Record Type", True),
    ("record_type_part", "Record Type (if Others)", False),
    ("record_date", "Record Date (YYYY-MM-DD)", True),
    ("pdf_link", "PDF Link", False),
    ("cabinet_number", "Cabinet Number", True),
    ("shelf_number", "Shelf Number", True),
    ("keywords", "Keywords", False),
    ("remarks", "Remarks", False),
    ("encoded_by", "Encoded By", False),
]

def _excel_styles():
    from openpyxl.styles import Font, PatternFill

    return (
        PatternFill("solid", fgColor="0033A0"),
        Font(color="FFFFFF", bold=True),
    )


def _field_keys(fields):
    return [field[0] for field in fields]


def _field_headers(fields):
    return [field[1] for field in fields]


def _required_keys(fields):
    return {field[0] for field in fields if field[2]}


def _normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return parse_document_date(value).strftime("%Y-%m-%d")
        except ValueError:
            return str(value).strip()
    return str(value).strip()


def _json_safe_import_data(data):
    if not data:
        return data
    safe = dict(data)
    for key in ("date_received", "date_accomp", "record_date"):
        value = safe.get(key)
        if isinstance(value, (date, datetime)):
            safe[key] = value.strftime("%Y-%m-%d")
    return safe


def _coerce_import_date(value, required=False):
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            raise ValueError("Date is required.")
        return None
    return parse_document_date(value).date()


def build_template_workbook(fields, sheet_name, notes):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    header_fill, header_font = _excel_styles()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    headers = _field_headers(fields)
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 18)

    instructions = workbook.create_sheet("Instructions")
    instructions["A1"] = notes
    instructions.column_dimensions["A"].width = 100
    return workbook


def workbook_to_bytes(workbook):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_dts_template_workbook():
    notes = (
        "Fill one row per document. Required columns must not be empty. "
        "Use valid values from the DTS forms (classification, doc type, status, etc.)."
    )
    return build_template_workbook(DTS_IMPORT_FIELDS, "dts_docs", notes)


def build_rmt_template_workbook():
    notes = (
        "Fill one row per record. Record numbers must be unique. "
        "Record type must match the RMT module options."
    )
    return build_template_workbook(RMT_IMPORT_FIELDS, "rmt_records", notes)


def _map_headers(header_row, fields):
    header_to_key = {label.lower(): key for key, label, _ in fields}
    mapping = {}
    for idx, header in enumerate(header_row, start=1):
        normalized = _normalize_cell(header).lower()
        if normalized in header_to_key:
            mapping[idx] = header_to_key[normalized]
    return mapping


def parse_workbook_rows(upload, fields):
    from openpyxl import load_workbook

    workbook = load_workbook(upload, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], ["The uploaded file is empty."]

    header_map = _map_headers(rows[0], fields)
    if not header_map:
        return [], ["Could not match any column headers. Use the provided template."]

    parsed_rows = []
    errors = []
    required = _required_keys(fields)

    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or all(_normalize_cell(cell) == "" for cell in row):
            continue

        record = {key: "" for key in _field_keys(fields)}
        for col_idx, key in header_map.items():
            if col_idx - 1 < len(row):
                record[key] = _normalize_cell(row[col_idx - 1])

        row_errors = []
        for key in required:
            if not record.get(key):
                row_errors.append(f"Missing required value for '{key}'.")

        parsed_rows.append({
            "row_number": row_idx,
            "data": record,
            "errors": row_errors,
            "valid": not row_errors,
        })

    if not parsed_rows:
        errors.append("No data rows found below the header row.")

    return parsed_rows, errors


def _parse_optional_date(value, label):
    if not value:
        return None, None
    try:
        return parse_document_date(value).date(), None
    except ValueError:
        return None, f"Invalid {label} date format. Use YYYY-MM-DD."


def validate_dts_row(record):
    data = dict(record)
    errors = []

    route_number = data.get("route_number", "").strip()
    if not route_number:
        errors.append("Route number is required.")
    elif DtsDoc.query.filter_by(route_number=route_number).first():
        errors.append(f"Route number '{route_number}' already exists in the database.")

    classification = normalize_classification(data.get("classification", "").strip())
    if classification not in CLASSIFICATIONS:
        errors.append("Classification must be Incoming or Outgoing.")

    doc_type, doc_type_part, doc_type_error = parse_doc_type_fields(
        data.get("doc_type"),
        data.get("doc_type_part"),
    )
    if doc_type_error:
        errors.append(doc_type_error)
    elif classification and not is_valid_doc_type(classification, doc_type):
        errors.append("Invalid document type for the selected classification.")

    action_needed, action_part, action_error = parse_action_fields(
        data.get("action_needed"),
        data.get("action_part"),
    )
    if action_error:
        errors.append(action_error)
    elif action_needed and classification and not is_valid_action_needed(classification, action_needed):
        errors.append("Invalid action needed for the selected classification.")

    origin, other_office, origin_error = resolve_origin_fields(
        classification,
        data.get("origin"),
        data.get("other_office"),
    )
    if origin_error:
        errors.append(origin_error)

    forwarded_to, forwarded_to_part, recipient_error = parse_recipient_fields(
        classification,
        data.get("forwarded_to"),
        data.get("forwarded_to_part"),
    )
    if recipient_error:
        errors.append(recipient_error)

    date_received, date_error = _parse_optional_date(data.get("date_received"), "date received")
    if date_error:
        errors.append(date_error)
    elif not date_received:
        errors.append("Date received is required.")

    date_accomp, accomp_error = _parse_optional_date(data.get("date_accomp"), "date accomplished")
    if accomp_error:
        errors.append(accomp_error)

    status = data.get("status", "").strip() or "Pending"
    if status not in STATUSES:
        errors.append(f"Status must be one of: {', '.join(STATUSES)}.")

    resp_unit = data.get("resp_unit", "").strip()
    if resp_unit and resp_unit not in list_department_names():
        errors.append("Invalid responsible unit.")

    if not data.get("subject", "").strip():
        errors.append("Subject is required.")

    if errors:
        return None, errors

    return {
        "route_number": route_number,
        "classification": classification,
        "doc_type": doc_type,
        "doc_type_part": doc_type_part,
        "date_received": date_received,
        "origin": origin,
        "other_office": other_office,
        "subject": data.get("subject", "").strip(),
        "doc_link": data.get("doc_link", "").strip() or None,
        "action_needed": action_needed,
        "action_part": action_part,
        "action_particulars": data.get("action_particulars", "").strip() or None,
        "resp_unit": resp_unit or None,
        "responsible_person": data.get("responsible_person", "").strip() or None,
        "action_provided": data.get("action_provided", "").strip() or None,
        "date_accomp": date_accomp,
        "status": status,
        "forwarded_to": forwarded_to,
        "forwarded_to_part": forwarded_to_part,
        "remarks": data.get("remarks", "").strip() or None,
    }, []


def preview_dts_import(upload):
    rows, file_errors = parse_workbook_rows(upload, DTS_IMPORT_FIELDS)
    if file_errors:
        return {"ok": False, "errors": file_errors, "rows": []}

    seen_routes = set()
    preview_rows = []
    valid_count = 0

    for item in rows:
        row_errors = list(item["errors"])
        route_number = item["data"].get("route_number", "").strip()
        if route_number:
            if route_number in seen_routes:
                row_errors.append(f"Duplicate route number '{route_number}' in file.")
            else:
                seen_routes.add(route_number)

        validated, validation_errors = validate_dts_row(item["data"])
        row_errors.extend(validation_errors)
        is_valid = not row_errors
        if is_valid:
            valid_count += 1

        preview_rows.append({
            "row_number": item["row_number"],
            "route_number": route_number,
            "subject": item["data"].get("subject", ""),
            "classification": item["data"].get("classification", ""),
            "status": item["data"].get("status", "") or "Pending",
            "errors": row_errors,
            "valid": is_valid,
            "data": _json_safe_import_data(validated),
        })

    return {
        "ok": True,
        "rows": preview_rows,
        "total": len(preview_rows),
        "valid_count": valid_count,
        "invalid_count": len(preview_rows) - valid_count,
    }


def commit_dts_import(preview_rows, actor_name, ip_address=None):
    created = 0
    for row in preview_rows:
        if not row.get("valid") or not row.get("data"):
            continue
        payload = dict(row["data"])
        payload["date_received"] = _coerce_import_date(payload.get("date_received"), required=True)
        payload["date_accomp"] = _coerce_import_date(payload.get("date_accomp"), required=False)
        doc = DtsDoc(**payload)
        db.session.add(doc)
        db.session.flush()
        log_history(doc, actor_name, ip_address)
        created += 1

    db.session.commit()
    return created


def validate_rmt_row(record):
    data = dict(record)
    errors = []

    record_number = data.get("record_number", "").strip()
    if not record_number:
        errors.append("Record number is required.")
    elif RmtRecord.query.filter_by(record_number=record_number).first():
        errors.append(f"Record number '{record_number}' already exists in the database.")

    title = data.get("title", "").strip()
    if not title:
        errors.append("Title is required.")

    record_type, record_type_part, type_error = parse_rmt_record_type_fields(
        data.get("record_type"),
        data.get("record_type_part"),
    )
    if type_error:
        errors.append(type_error)
    elif record_type and not is_valid_rmt_record_type(record_type):
        errors.append("Invalid record type.")

    cabinet_number = data.get("cabinet_number", "").strip()
    shelf_number = data.get("shelf_number", "").strip()
    if not cabinet_number:
        errors.append("Cabinet number is required.")
    if not shelf_number:
        errors.append("Shelf number is required.")

    record_date, date_error = _parse_optional_date(data.get("record_date"), "record")
    if date_error:
        errors.append(date_error)
    elif not record_date:
        errors.append("Record date is required.")

    if errors:
        return None, errors

    return {
        "record_number": record_number,
        "title": title,
        "record_type": record_type,
        "record_type_part": record_type_part,
        "record_date": record_date,
        "pdf_link": data.get("pdf_link", "").strip() or None,
        "cabinet_number": cabinet_number,
        "shelf_number": shelf_number,
        "keywords": data.get("keywords", "").strip() or None,
        "remarks": data.get("remarks", "").strip() or None,
        "encoded_by": data.get("encoded_by", "").strip() or None,
    }, []


def preview_rmt_import(upload):
    rows, file_errors = parse_workbook_rows(upload, RMT_IMPORT_FIELDS)
    if file_errors:
        return {"ok": False, "errors": file_errors, "rows": []}

    seen_numbers = set()
    preview_rows = []
    valid_count = 0

    for item in rows:
        row_errors = list(item["errors"])
        record_number = item["data"].get("record_number", "").strip()
        if record_number:
            if record_number in seen_numbers:
                row_errors.append(f"Duplicate record number '{record_number}' in file.")
            else:
                seen_numbers.add(record_number)

        validated, validation_errors = validate_rmt_row(item["data"])
        row_errors.extend(validation_errors)
        is_valid = not row_errors
        if is_valid:
            valid_count += 1

        preview_rows.append({
            "row_number": item["row_number"],
            "record_number": record_number,
            "title": item["data"].get("title", ""),
            "record_type": item["data"].get("record_type", ""),
            "errors": row_errors,
            "valid": is_valid,
            "data": _json_safe_import_data(validated),
        })

    return {
        "ok": True,
        "rows": preview_rows,
        "total": len(preview_rows),
        "valid_count": valid_count,
        "invalid_count": len(preview_rows) - valid_count,
    }


def commit_rmt_import(preview_rows):
    created = 0
    for row in preview_rows:
        if not row.get("valid") or not row.get("data"):
            continue
        payload = dict(row["data"])
        payload["record_date"] = _coerce_import_date(payload.get("record_date"), required=True)
        db.session.add(RmtRecord(**payload))
        created += 1

    db.session.commit()
    return created


def purge_module_data(module_key):
    module_key = (module_key or "").strip().lower()
    if module_key == "dts":
        history_count = DtsHistory.query.count()
        docs_count = DtsDoc.query.count()
        DtsHistory.query.delete()
        DtsDoc.query.delete()
        db.session.commit()
        return {
            "module": "DTS",
            "removed": {"dts_history": history_count, "dts_docs": docs_count},
        }
    if module_key == "rmt":
        records_count = RmtRecord.query.count()
        RmtRecord.query.delete()
        db.session.commit()
        return {"module": "RMT", "removed": {"rmt_records": records_count}}
    if module_key == "pmis":
        return {
            "module": "PMIS",
            "removed": {},
            "message": "PMIS has no database tables yet. Nothing to purge.",
        }
    raise ValueError("Unknown module.")


def module_record_counts():
    return {
        "dts_docs": DtsDoc.query.count(),
        "dts_history": DtsHistory.query.count(),
        "rmt_records": RmtRecord.query.count(),
        "pmis_records": 0,
    }
